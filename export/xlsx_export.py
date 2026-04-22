import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

_BLUE_FILL = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
_LIGHT_BLUE_FILL = PatternFill(start_color="E8EFFF", end_color="E8EFFF", fill_type="solid")
_ALT_ROW_FILL = PatternFill(start_color="F7F9FF", end_color="F7F9FF", fill_type="solid")
_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Arial", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="CDD5E8"),
    right=Side(style="thin", color="CDD5E8"),
    top=Side(style="thin", color="CDD5E8"),
    bottom=Side(style="thin", color="CDD5E8"),
)


def _style_header_row(ws, row_num: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _BLUE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _style_body_rows(ws, start_row: int, end_row: int, col_count: int):
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _THIN_BORDER
            if (row - start_row) % 2 == 1:
                cell.fill = _ALT_ROW_FILL


def _auto_column_width(ws, col_count: int, max_width: int = 50):
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max_len + 4


def generate_xlsx(data: dict) -> bytes:
    wb = Workbook()

    # Sheet 1: Incident Summary
    ws_summary = wb.active
    ws_summary.title = "Incident Summary"
    stats = data.get("stats", {})

    ws_summary.cell(row=1, column=1, value="Metric")
    ws_summary.cell(row=1, column=2, value="Value")
    _style_header_row(ws_summary, 1, 2)

    summary_rows = [
        ("Total Incidents", stats.get("total", 0)),
    ]
    for severity, count in stats.get("by_severity", {}).items():
        summary_rows.append((f"Severity: {severity}", count))
    for status, count in stats.get("by_status", {}).items():
        summary_rows.append((f"Status: {status}", count))

    for i, (metric, value) in enumerate(summary_rows, start=2):
        ws_summary.cell(row=i, column=1, value=metric)
        ws_summary.cell(row=i, column=2, value=value)
    _style_body_rows(ws_summary, 2, 1 + len(summary_rows), 2)
    _auto_column_width(ws_summary, 2)

    # Sheet 2: Incident Details
    ws_details = wb.create_sheet("Incident Details")
    headers = ["Issue Key", "Summary", "Status", "Priority", "Severity",
               "Labels", "Created", "Resolved", "Assignee", "Close Justification"]
    for ci, h in enumerate(headers, start=1):
        ws_details.cell(row=1, column=ci, value=h)
    _style_header_row(ws_details, 1, len(headers))

    incidents = data.get("incidents", [])
    for ri, inc in enumerate(incidents, start=2):
        ws_details.cell(row=ri, column=1, value=inc.get("key", ""))
        ws_details.cell(row=ri, column=2, value=inc.get("summary", ""))
        ws_details.cell(row=ri, column=3, value=inc.get("status", ""))
        ws_details.cell(row=ri, column=4, value=inc.get("priority", ""))
        ws_details.cell(row=ri, column=5, value=inc.get("severity", ""))
        ws_details.cell(row=ri, column=6, value=", ".join(inc.get("labels", [])))
        ws_details.cell(row=ri, column=7, value=inc.get("created", ""))
        ws_details.cell(row=ri, column=8, value=inc.get("resolved", ""))
        ws_details.cell(row=ri, column=9, value=inc.get("assignee", ""))
        ws_details.cell(row=ri, column=10, value=inc.get("close_justification", ""))

    if incidents:
        _style_body_rows(ws_details, 2, 1 + len(incidents), len(headers))
    _auto_column_width(ws_details, len(headers))

    # Sheet 3: Severity Breakdown
    ws_severity = wb.create_sheet("Severity Breakdown")
    ws_severity.cell(row=1, column=1, value="Severity")
    ws_severity.cell(row=1, column=2, value="Count")
    _style_header_row(ws_severity, 1, 2)

    by_severity = stats.get("by_severity", {})
    for i, (sev, count) in enumerate(by_severity.items(), start=2):
        ws_severity.cell(row=i, column=1, value=sev)
        ws_severity.cell(row=i, column=2, value=count)
    if by_severity:
        _style_body_rows(ws_severity, 2, 1 + len(by_severity), 2)
    _auto_column_width(ws_severity, 2)

    # Sheet 4: Monthly Trend
    ws_trend = wb.create_sheet("Monthly Trend")
    ws_trend.cell(row=1, column=1, value="Month")
    ws_trend.cell(row=1, column=2, value="Incident Count")
    _style_header_row(ws_trend, 1, 2)

    monthly = stats.get("monthly_trend", {})
    for i, (month, count) in enumerate(monthly.items(), start=2):
        ws_trend.cell(row=i, column=1, value=month)
        ws_trend.cell(row=i, column=2, value=count)
    if monthly:
        _style_body_rows(ws_trend, 2, 1 + len(monthly), 2)
    _auto_column_width(ws_trend, 2)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
